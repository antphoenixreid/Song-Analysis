from pathlib import Path
import os
from concurrent.futures import ProcessPoolExecutor

import numpy as np
import pandas as pd
import librosa
from pydub import AudioSegment

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from audio_features.audio_signal import AudioSignal
from audio_features.time_features import TimeFeatures
from audio_features.frequency_features import FrequencyFeatures
from audio_features.chromagram_features import ChromagramFeatures
from audio_features.tempogram_features import TempogramFeatures
from audio_features.mfcc_features import MFCCFeatures


TARGET_SR = 22050
SUPPORTED_EXTENSIONS = {".wav", ".mp3", ".flac", ".ogg", ".m4a"}


def load_audio(audio_path, N=2048, H=512):
    audio_path = Path(audio_path)
    try:
        y, sr = librosa.load(audio_path, sr=TARGET_SR, mono=True)
        backend = "librosa"
    except Exception as e:
        print(f" -> [Backend Fallback] {audio_path.name}: {e}")
        audio = AudioSegment.from_file(audio_path)
        audio = audio.set_frame_rate(TARGET_SR).set_channels(1)
        y = np.array(audio.get_array_of_samples(), dtype=np.float32)
        y /= float(2 ** (8 * audio.sample_width - 1))
        sr = TARGET_SR
        backend = "pydub"

    sig = AudioSignal(audio_path=str(audio_path), signal=y, sr=sr, N=N, H=H)
    return sig, backend


def safe_mean(x):
    x = np.asarray(x)
    return float(np.mean(x)) if x.size else 0.0


def safe_std(x):
    x = np.asarray(x)
    return float(np.std(x)) if x.size else 0.0


def safe_first(x, default=0.0):
    x = np.asarray(x)
    return float(x.flat[0]) if x.size else default


def status_from_values(row, keys):
    vals = [row.get(k, None) for k in keys]
    if any(v is None for v in vals):
        return "MISSING"
    if any(isinstance(v, (int, float, np.integer, np.floating)) and not np.isfinite(v) for v in vals):
        return "NAN"
    if any(v == 0 for v in vals):
        return "ZERO"
    if any(v == 1 for v in vals):
        return "ONE"
    return "OK"

def safe_circular_mean(x):
    """
    Computes the circular mean for phases in [0, 1).
    Returns a value normalized back into [0, 1).
    """
    arr = np.asanyarray(x, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return 0.0

    angles = 2.0 * np.pi * arr
    sin_mean = np.mean(np.sin(angles))
    cos_mean = np.mean(np.cos(angles))

    R = np.hypot(sin_mean, cos_mean)
    if R < 1e-6:
        return 0.5

    mean_angle = np.arctan2(sin_mean, cos_mean)
    return float((mean_angle / (2.0 * np.pi)) % 1.0)

def safe_circular_std(x):
    """
    Computes the circular standard deviation (Yamartino estimator) for phases in [0, 1).
    Returns standard deviation mapped back into [0, 1).
    """
    arr = np.asanyarray(x, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return 0.0

    angles = 2.0 * np.pi * arr
    R = np.hypot(np.mean(np.sin(angles)), np.mean(np.cos(angles)))

    if R < 1e-6:
        return 0.5

    circ_std = np.sqrt(max(-2.0 * np.log(R), 0.0))
    return float(np.clip(circ_std / (2.0 * np.pi), 0.0, 0.5))

def export_time_debug_row(audio_path, sig):
    tf = TimeFeatures(sig)

    silence_stats = tf._silence_duration(db_threshold=-60.0)
    rhythm_stats = tf._rhythmic_stability()
    hjorth = tf._hjorth_parameters()
    ioi_metrics = tf._ioi_stats()

    rms_env = tf._rms_envelope()
    st_energy = tf._short_time_energy()
    peak_amp = tf._peak_amplitude()
    zcr = tf._zero_crossing_rate()

    row = {
        "file_name": Path(audio_path).name,
        "file_path": str(audio_path),
        "backend": getattr(sig, "backend", ""),
        "global_loudness_dB": tf._global_loudness_dB(),
        "rms_envelope_mean": safe_mean(rms_env),
        "rms_envelope_std": safe_std(rms_env),
        "short_time_energy_mean": safe_mean(st_energy),
        "peak_amplitude_mean": safe_mean(peak_amp),
        "dynamic_range": tf._dynamic_range(),
        "attack_time": tf._attack_time(),
        "attack_slope": tf._attack_slope(),
        "decay_slope": tf._decay_slope(),
        "energy_variance": tf._energy_variance(),
        "energy_modulation_rate": tf._energy_modulation_rate(),
        "zero_crossing_rate_mean": safe_mean(zcr),
        "zcr_variance": tf._zcr_variance(),
        "voiced_ratio": tf._voiced_ratio(),
        "unvoiced_ratio": tf._unvoiced_ratio(),
        "transient_rate": tf._transient_rate(),
        "transient_counts": tf._transient_counts(),
        "onset_rate": tf._onset_rate(),
        "mean_ioi": ioi_metrics[0],
        "std_ioi": ioi_metrics[1],
        "cv_ioi": ioi_metrics[2],
        "tempo_from_onset_ac": tf._tempo_from_onset_ac(),
        "pulse_clarity_ac": tf._pulse_clarity_ac(),
        "silence_ratio": tf._silence_ratio(db_threshold=-60.0),
        "silence_total_sec": silence_stats["total"],
        "silence_max_sec": silence_stats["max"],
        "silence_mean_sec": silence_stats["mean"],
        "silence_count": silence_stats["count"],
        "rhythmic_tempo_var": rhythm_stats["tempo_var"],
        "rhythmic_stability_exp": rhythm_stats["stability_exp"],
        "rhythmic_stability_cv": rhythm_stats["stability_cv"],
        "lz_complexity": tf._lz_complexity(),
        "higuchi_fd": tf._higuchi_fd(),
        "hjorth_activity": hjorth["activity"],
        "hjorth_mobility": hjorth["mobility"],
        "hjorth_complexity": hjorth["complexity"],
        "spotify_loudness_db": tf._spotify_loudness(active_only=True),
        "spotify_energy": tf._spotify_energy(active_only=True),
        "spotify_speechiness": tf._spotify_speechiness(),
        "spotify_acousticness": tf._spotify_acousticness(),
        "spotify_danceability": tf._spotify_danceability(),
        "spotify_tempo": tf._spotify_tempo(),
        "spotify_liveness": tf._spotify_liveness(),
        "spotify_instrumentalness": tf._spotify_instrumentalness(),
        "spotify_time_signature": tf._spotify_time_signature(),
    }

    row["status"] = status_from_values(row, ["attack_slope", "transient_counts", "spotify_energy"])
    row["notes"] = ""
    return row


def export_freq_debug_row(audio_path, sig):
    ff = FrequencyFeatures(sig)

    pc = ff._pitch_class_profile(use_power=True)
    key_mode = ff._estimate_key_mode_freq()
    spotify = ff.spotify_audio_features()

    frame_energy = ff._frame_energy()
    frame_energy_db = ff._frame_energy_db()

    row = {
        "file_name": Path(audio_path).name,
        "file_path": str(audio_path),
        "frame_energy_mean": safe_mean(frame_energy),
        "frame_energy_db_mean": safe_mean(frame_energy_db),
        "dynamic_range": ff._dynamic_range(),
        "spectral_centroid_mean": safe_mean(ff._spectral_centroid()),
        "spectral_bandwidth_mean": safe_mean(ff._spectral_bandwidth()),
        "spectral_rolloff_mean": safe_mean(ff._spectral_rolloff()),
        "spectral_slope_mean": safe_mean(ff._spectral_slope()),
        "spectral_skewness_mean": safe_mean(ff._spectral_skewness()),
        "spectral_kurtosis_mean": safe_mean(ff._spectral_kurtosis()),
        "spectral_flatness_mean": safe_mean(ff._spectral_flatness()),
        "spectral_entropy_mean": safe_mean(ff._spectral_entropy()),
        "spectral_flux_mean": safe_mean(ff._spectral_flux()),
        "pulse_clarity_ac": ff._pulse_clarity_ac(),
        "beat_periodicity": ff._beat_periodicity(),
        "transient_counts": ff._transient_counts(),
        "transient_rate": ff._transient_rate(),
        "phase_congruency_mean": safe_mean(ff._phase_congruency()),
        "phase_coherence_time_mean": safe_mean(ff._phase_coherence_time()),
        "harmonic_ratio_mean": safe_mean(ff._harmonic_ratio()) if ff._harmonic_ratio().size else 0.0,
        "inharmonicity_mean": safe_mean(ff._inharmonicity()) if ff._inharmonicity().size else 0.0,
        "hnr_mean": safe_mean(ff._hnr(f0_hz=ff._fundamental_freq_estimate())) if ff._fundamental_freq_estimate().size else 0.0,
        "sub_band_entropy_mean": safe_mean(ff._sub_band_entropy()),
        "sub_band_flatness_mean": safe_mean(ff._sub_band_flatness()),
        "sub_band_centroid_mean": safe_mean(ff._sub_band_centroid()),
        "sub_band_low_high_ratio_mean": safe_mean(ff._sub_band_low_high_ratio()),
        "pitch_class_0": float(pc[0]),
        "pitch_class_1": float(pc[1]),
        "pitch_class_2": float(pc[2]),
        "pitch_class_3": float(pc[3]),
        "pitch_class_4": float(pc[4]),
        "pitch_class_5": float(pc[5]),
        "pitch_class_6": float(pc[6]),
        "pitch_class_7": float(pc[7]),
        "pitch_class_8": float(pc[8]),
        "pitch_class_9": float(pc[9]),
        "pitch_class_10": float(pc[10]),
        "pitch_class_11": float(pc[11]),
        "key": key_mode["tonic"],
        "mode": key_mode["mode"],
        "key_score": key_mode["score"],
        "spotify_loudness_db": spotify["loudness_db"],
        "spotify_energy": spotify["energy"],
        "spotify_speechiness": spotify["speechiness"],
        "spotify_acousticness": spotify["acousticness"],
        "spotify_danceability": spotify["danceability"],
        "spotify_valence": spotify["valence"],
        "spotify_tempo_bpm": spotify["tempo_bpm"],
        "spotify_liveness": spotify["liveness"],
        "spotify_instrumentalness": spotify["instrumentalness"],
        "spotify_key": spotify["key"],
        "spotify_mode": spotify["mode"],
        "spotify_time_signature": spotify["time_signature"],
        "spotify_fused": spotify["spotify_fused"],
    }

    row["status"] = status_from_values(row, ["frame_energy_mean", "key_score", "spotify_energy"])
    row["notes"] = ""
    return row


def export_chroma_debug_row(audio_path, sig):
    cf = ChromagramFeatures(sig)

    mean_chroma = cf._mean_chroma(normalize=True, use_db=False)
    pitch_profile = cf._pitch_class_profile(normalize=True, use_db=False)
    key_res = cf._key_estimation(normalize=True, use_db=False, method="cosine")
    mode_res = cf._mode_classification(normalize=True, use_db=False, method="cosine")
    tonal_clarity = cf._tonal_clarity(normalize=True, use_db=False, method="cosine")
    harm_entropy = cf._harmonic_entropy(normalize=True, use_db=False, method="cosine")
    cons_dis = cf._consonance_dissonance(normalize=True, use_db=False, method="cosine")
    chord_det = cf._chord_detection(normalize=True, use_db=False, method="cosine")
    chord_prog = cf._chord_progression_mapping(normalize=True, use_db=False, method="cosine")
    harm_rhythm = cf._harmonic_rhythm(normalize=True, use_db=False, method="cosine")
    root_motion = cf._root_motion_analysis(normalize=True, use_db=False, method="cosine")
    tsi = cf._tonal_stability_index(normalize=True, use_db=False, method="cosine")
    acf = cf._chroma_autocorrelation(lag_max=64, normalize=True, use_db=False)
    variability = cf._chroma_variability(normalize=True, use_db=False)
    smoothness = cf._chroma_smoothness(normalize=True, use_db=False, metric="l2")
    dominant = cf._dominant_pitch_track(normalize=True, use_db=False)
    tuning = cf._tuning_deviation_detection(top_k=20)
    spotify = cf.spotify_audio_features(normalize=True, use_db=False, method="cosine")

    row = {
        "file_name": Path(audio_path).name,
        "file_path": str(audio_path),
        "mean_chroma_0": float(mean_chroma[0]),
        "mean_chroma_1": float(mean_chroma[1]),
        "mean_chroma_2": float(mean_chroma[2]),
        "mean_chroma_3": float(mean_chroma[3]),
        "mean_chroma_4": float(mean_chroma[4]),
        "mean_chroma_5": float(mean_chroma[5]),
        "mean_chroma_6": float(mean_chroma[6]),
        "mean_chroma_7": float(mean_chroma[7]),
        "mean_chroma_8": float(mean_chroma[8]),
        "mean_chroma_9": float(mean_chroma[9]),
        "mean_chroma_10": float(mean_chroma[10]),
        "mean_chroma_11": float(mean_chroma[11]),
        "pitch_profile_0": float(pitch_profile[0].mean()),
        "pitch_profile_1": float(pitch_profile[1].mean()),
        "pitch_profile_2": float(pitch_profile[2].mean()),
        "pitch_profile_3": float(pitch_profile[3].mean()),
        "pitch_profile_4": float(pitch_profile[4].mean()),
        "pitch_profile_5": float(pitch_profile[5].mean()),
        "pitch_profile_6": float(pitch_profile[6].mean()),
        "pitch_profile_7": float(pitch_profile[7].mean()),
        "pitch_profile_8": float(pitch_profile[8].mean()),
        "pitch_profile_9": float(pitch_profile[9].mean()),
        "pitch_profile_10": float(pitch_profile[10].mean()),
        "pitch_profile_11": float(pitch_profile[11].mean()),
        "key_idx": key_res["key_idx"],
        "tonic": key_res["tonic"],
        "mode": key_res["mode"],
        "key_score": key_res["score"],
        "mode_score_major": mode_res["score_major"],
        "mode_score_minor": mode_res["score_minor"],
        "mode_delta_score": mode_res["delta_score"],
        "tonal_clarity": tonal_clarity["tonal_clarity"],
        "harmonic_entropy": harm_entropy["harmonic_entropy"],
        "raw_entropy": harm_entropy["raw_entropy"],
        "consonance": cons_dis["consonance"],
        "dissonance": cons_dis["dissonance"],
        "best_chord_idx": int(chord_det["chord_idx"][0]) if len(chord_det["chord_idx"]) > 0 else -1,
        "best_chord_score_first": float(chord_det["best_scores"][0]) if len(chord_det["best_scores"]) > 0 else 0.0,
        "harmonic_change_rate": harm_rhythm["change_rate"],
        "avg_chord_duration_sec": harm_rhythm["avg_duration_sec"],
        "num_changes": harm_rhythm["num_changes"],
        "avg_root_motion": root_motion["avg_motion"],
        "tonal_stability_index": tsi,
        "chroma_acf_lag1": float(acf["acf_mean"][1]) if acf["acf_mean"].size > 1 else 0.0,
        "chroma_variability_mean": variability["mean_variability"],
        "chroma_variability_median": variability["median_variability"],
        "chroma_smoothness": smoothness["smoothness"],
        "dominant_pitch_mean": float(dominant.mean()) if dominant.size else 0.0,
        "track_cents": tuning["track_cents"],
        "mean_abs_cents": tuning.get("mean_abs_cents", 0.0),
        "spotify_energy": spotify["energy"],
        "spotify_speechiness": spotify["speechiness"],
        "spotify_acousticness": spotify["acousticness"],
        "spotify_danceability": spotify["danceability"],
        "spotify_valence": spotify["valence"],
        "spotify_tempo": spotify["tempo"],
        "spotify_instrumentalness": spotify["instrumentalness"],
        "spotify_key": spotify["key"],
        "spotify_mode": spotify["mode"],
        "spotify_time_signature": spotify["time_signature"],
        "chroma_entropy": cf._chroma_entropy(normalize=True, use_db=False),
        "chroma_flux_mean": cf._chroma_flux_mean(normalize=True, use_db=False),
        "chroma_flux_variance": cf._chroma_flux_variance(normalize=True, use_db=False),
        "harmonic_template_fit": cf._harmonic_template_fit(normalize=True, use_db=False, method="cosine")["best_score"],
        "pitch_class_peakedness": cf._pitch_class_peakedness(normalize=True, use_db=False),
        "energy_chroma": cf._energy_chroma(normalize=True, use_db=False),
        "speechiness_chroma": cf._speechiness_chroma(normalize=True, use_db=False),
        "acousticness_chroma": cf._acousticness_chroma(normalize=True, use_db=False),
        "danceability_chroma": cf._danceability_chroma(normalize=True, use_db=False),
        "valence_chroma": cf._valence_chroma(normalize=True, use_db=False),
        "tempo_chroma": cf._tempo_chroma(normalize=True, use_db=False),
        "instrumentalness_chroma": cf._instrumentalness_chroma(normalize=True, use_db=False),
    }

    row["status"] = status_from_values(row, ["energy_chroma", "harmonic_entropy", "key_score"])
    row["notes"] = ""
    return row


def export_tempo_debug_row(audio_path, sig):
    tf = TempogramFeatures(sig)

    onset = tf._onset_strength(max_size=1, detrend=False, aggregate=np.mean, smooth=False)
    onset_env = onset["onset_env"]
    temp_ac = tf._tempogram_autocorr(win_length=sig.N, center=True, norm_sum=True)
    temp_f = tf._tempogram_fourier(win_length=sig.N, center=True, window="hann")
    g = tf._global_bpm(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    l = tf._local_bpm_curve(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    pulse = tf._pulse_clarity(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    stability = tf._tempo_stability_index(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    variation = tf._tempo_variation_curve(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    multi = tf._multi_periodic_structure(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    swing = tf._swing_ratio(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    centroid = tf._tempo_spectral_centroid(bpm_min=40.0, bpm_max=240.0)
    bandwidth = tf._tempo_bandwidth(bpm_min=40.0, bpm_max=240.0)
    skewness = tf._tempo_skewness(bpm_min=40.0, bpm_max=240.0)
    kurtosis = tf._tempo_kurtosis(bpm_min=40.0, bpm_max=240.0)
    beat_pos = tf._beat_position(beat_frames=None, beat_times=None, mode="phase")
    beat_hist = tf._beat_alignment_histogram(beat_frames=None, beat_times=None, n_bins=16, normalize=True)
    ibi_var = tf._interbeat_interval_variance(beat_frames=None, beat_times=None, normalize=True)
    sync_offset = tf._beat_sync_offset(beat_frames=None, beat_times=None, event_frames=None, event_times=None, absolute=True)
    rhythmic_energy = tf._beat_periodic_energy(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    dance = tf._danceability_tempogram(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    valence = tf._valence_tempogram(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    liveness = tf._liveness_tempogram(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    mode = tf._mode_tempogram(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    tsig = tf._time_signature_tempogram(bpm_min=40.0, bpm_max=240.0, norm_sum=True)
    spotify = tf.spotify_audio_features(beat_times=None, beat_frames=None, bpm_min=40.0, bpm_max=240.0, norm_sum=True)

    row = {
        "file_name": Path(audio_path).name,
        "file_path": str(audio_path),
        "onset_mean": safe_mean(onset_env),
        "onset_sum": float(onset_env.sum()) if onset_env.size else 0.0,
        "onset_energy": tf._onset_energy(normalize=False, max_size=1, detrend=False),
        "onset_energy_mean": tf._onset_energy(normalize=True, max_size=1, detrend=False),
        "transient_peak_count": tf._transient_curve(smooth=False, smooth_width=5, normalize=True)["peak_count"],
        "transient_mean_slope": tf._transient_curve(smooth=False, smooth_width=5, normalize=True)["mean_slope"],
        "periodicity": tf._envelope_periodicity(lag_max=None, normalize=True, method="autocorr")["periodicity"],
        "best_lag": tf._envelope_periodicity(lag_max=None, normalize=True, method="autocorr").get("best_lag", 0),
        "global_bpm": g["bpm"],
        "global_strength": g["strength"],
        "global_lag": g["lag"],
        "pulse_clarity": pulse["clarity"],
        "pulse_best_peak": pulse["best_peak"],
        "pulse_runner_up": pulse["runner_up"],
        "tempo_stability": stability["stability"],
        "tempo_mean_bpm": stability["mean_bpm"],
        "tempo_std_bpm": stability["std_bpm"],
        "tempo_variation_mean": float(variation["curve"].mean()) if variation["curve"].size else 0.0,
        "tempo_variation_max": float(variation["curve"].max()) if variation["curve"].size else 0.0,
        "multi_periodic_score": multi["score"],
        "primary_bpm": multi["primary_bpm"],
        "half_bpm": multi["half_bpm"],
        "double_bpm": multi["double_bpm"],
        "swing_ratio": swing["ratio"],
        "swing_symmetry": swing["symmetry"],
        "tempo_centroid": centroid["centroid"],
        "tempo_bandwidth": bandwidth["bandwidth"],
        "tempo_skewness": skewness["skewness"],
        "tempo_kurtosis": kurtosis["kurtosis"],
        "beat_phase_mean": safe_circular_mean(beat_pos["beat_position"]),
        "beat_phase_std": safe_circular_std(beat_pos["beat_position"]),
        "beat_hist_peak_bin": beat_hist["peak_bin"],
        "ibi_variance": ibi_var,
        "rhythmic_energy_mean": safe_mean(rhythmic_energy),
        "rhythmic_energy_std": safe_std(rhythmic_energy),
        "sync_mean_offset": sync_offset.get("mean_offset", 0.0),
        "sync_mean_abs_offset": sync_offset.get("mean_abs_offset", 0.0),
        "sync_mean_offset_norm": sync_offset.get("mean_offset_norm", 0.0),
        "sync_mean_abs_offset_norm": sync_offset.get("mean_abs_offset_norm", 0.0),
        "danceability_tempogram": dance,
        "valence_tempogram": valence,
        "liveness_tempogram": liveness,
        "mode_tempogram": mode["mode"],
        "mode_major_score": mode["score_major"],
        "mode_minor_score": mode["score_minor"],
        "mode_delta_score": mode["delta_score"],
        "time_signature_tempogram": tsig["time_signature"],
        "time_signature_confidence": tsig["confidence"],
        "time_signature_primary_bpm": tsig["primary_bpm"],
        "time_signature_structure_score": tsig["structure_score"],
        "spotify_loudness_per_beat_mean": float(spotify["loudness_per_beat"].mean()) if len(spotify["loudness_per_beat"]) else 0.0,
        "spotify_danceability": spotify["danceability"],
        "spotify_valence": spotify["valence"],
        "spotify_liveness": spotify["liveness"],
        "spotify_mode": spotify["mode"],
        "spotify_time_signature": spotify["time_signature"],
        "tempogram_autocorr_shape_0": int(temp_ac["tempogram"].shape[0]),
        "tempogram_autocorr_shape_1": int(temp_ac["tempogram"].shape[1]),
        "tempogram_fourier_shape_0": int(temp_f["tempogram"].shape[0]),
        "tempogram_fourier_shape_1": int(temp_f["tempogram"].shape[1]),
    }

    row["status"] = status_from_values(row, ["global_bpm", "primary_bpm", "beat_phase_mean"])
    row["notes"] = ""
    return row


def export_mfcc_debug_row(audio_path, sig):
    mf = MFCCFeatures(sig, n_mfcc=13, n_mels=40, compute=True)

    coeff_mean = mf._mfcc_mean()
    coeff_var = mf._mfcc_variance()
    coeff_skew = mf._mfcc_skewness()
    coeff_kurt = mf._mfcc_kurtosis(excess=True)
    delta = mf._mfcc_delta()
    delta2 = mf._mfcc_delta2()
    stability = mf._mfcc_temporal_stability()
    acf = mf._mfcc_autocorrelation(max_lag=64, normalize=True)

    loudness = mf._loudness_mfcc()
    energy = mf._energy_mfcc()
    speechiness = mf._speechiness_mfcc()
    acousticness = mf._acousticness_mfcc()
    valence = mf._valence_mfcc()
    liveness = mf._liveness_mfcc()
    instrumentalness = mf._instrumentalness_mfcc()
    spotify = mf.spotify_audio_features()

    row = {
        "file_name": Path(audio_path).name,
        "file_path": str(audio_path),
        "mfcc_energy": energy,
        "mfcc_loudness": loudness,
        "mfcc_speechiness": speechiness,
        "mfcc_acousticness": acousticness,
        "mfcc_valence": valence,
        "mfcc_liveness": liveness,
        "mfcc_instrumentalness": instrumentalness,
        "mfcc_smoothness": mf._mfcc_smoothness(normalize=True),
        "mfcc_entropy": mf._mfcc_entropy(normalize=True),
        "mfcc_flux": mf._mfcc_flux(normalize=True),
        "mfcc_high_order_energy": mf._mfcc_high_order_energy(start_coeff=6, normalize=True, order="l2"),
        "mfcc_high_order_variance": mf._mfcc_high_order_variance(start_coeff=6, normalize=True),
        "mfcc_transient_roughness": mf._mfcc_transient_roughness(width=9, normalize=True, mode="interp"),
        "mfcc_attack_smoothness": mf._mfcc_attack_smoothness(attack_frames=None, normalize=True),
        "mfcc_sustain_stability": mf._mfcc_sustain_stability(attack_frames=None, sustain_frames=None, normalize=True),
        "spotify_loudness": spotify["loudness"],
        "spotify_energy": spotify["energy"],
        "spotify_speechiness": spotify["speechiness"],
        "spotify_acousticness": spotify["acousticness"],
        "spotify_valence": spotify["valence"],
        "spotify_liveness": spotify["liveness"],
        "spotify_instrumentalness": spotify["instrumentalness"],
    }

    for i in range(len(coeff_mean)):
        row[f"mfcc_mean_{i}"] = float(coeff_mean[i])
        row[f"mfcc_var_{i}"] = float(coeff_var[i])
        row[f"mfcc_skew_{i}"] = float(coeff_skew[i])
        row[f"mfcc_kurt_{i}"] = float(coeff_kurt[i])
        row[f"mfcc_stability_{i}"] = float(stability[i])

    for i in range(delta.shape[0]):
        row[f"mfcc_delta_mean_{i}"] = float(np.mean(delta[i]))
        row[f"mfcc_delta2_mean_{i}"] = float(np.mean(delta2[i]))

    row["mfcc_acf_lag1"] = float(acf["acf"][0, 1]) if acf["acf"].shape[1] > 1 else 0.0
    row["mfcc_acf_lag2"] = float(acf["acf"][0, 2]) if acf["acf"].shape[1] > 2 else 0.0
    row["status"] = status_from_values(row, ["mfcc_energy", "mfcc_attack_smoothness", "mfcc_entropy"])
    row["notes"] = ""
    return row


def stylize_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    header_fill = PatternFill(start_color="2A3439", end_color="2A3439", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    border_side = Side(style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            ws.views.sheetView[0].showGridLines = True
        except Exception:
            pass

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(cell.value, (int, np.integer)):
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    wb.save(file_path)


def build_index_sheet(sheets_data):
    rows = []
    for sheet_name, df in sheets_data.items():
        rows.append({
            "sheet_name": sheet_name,
            "rows": len(df),
            "columns": len(df.columns),
            "status_counts": df["status"].value_counts().to_dict() if "status" in df.columns else {},
        })
    return pd.DataFrame(rows)


def analyze_audio_to_excel_debug(audio_path, output_excel_path):
    print(f"Processing: {Path(audio_path).name}")
    sig, backend = load_audio(audio_path)
    sig.backend = backend

    time_row = export_time_debug_row(audio_path, sig)
    freq_row = export_freq_debug_row(audio_path, sig)
    chroma_row = export_chroma_debug_row(audio_path, sig)
    tempo_row = export_tempo_debug_row(audio_path, sig)
    mfcc_row = export_mfcc_debug_row(audio_path, sig)

    sheets_data = {
        "Time": pd.DataFrame([time_row]),
        "Frequency": pd.DataFrame([freq_row]),
        "Chroma": pd.DataFrame([chroma_row]),
        "Tempo": pd.DataFrame([tempo_row]),
        "MFCC": pd.DataFrame([mfcc_row]),
    }
    sheets_data["Index"] = build_index_sheet(sheets_data)

    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        sheets_data["Index"].to_excel(writer, sheet_name="Index", index=False)
        for sheet_name in ["Time", "Frequency", "Chroma", "Tempo", "MFCC"]:
            sheets_data[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    stylize_excel_file(output_excel_path)
    print(f"Successfully generated debug workbook at: {output_excel_path}")
    return sheets_data


def process_audio_file(audio_path):
    audio_path = Path(audio_path)
    output_excel_path = audio_path.with_name(f"{audio_path.stem}_debug_features.xlsx")
    try:
        sig, _ = load_audio(str(audio_path))
        if hasattr(sig, "is_valid") and not sig.is_valid:
            print(f" [WARNING] Skipping {audio_path.name}: invalid audio.")
            return None

        return {
            "Time": pd.DataFrame([export_time_debug_row(str(audio_path), sig)]),
            "Frequency": pd.DataFrame([export_freq_debug_row(str(audio_path), sig)]),
            "Chroma": pd.DataFrame([export_chroma_debug_row(str(audio_path), sig)]),
            "Tempo": pd.DataFrame([export_tempo_debug_row(str(audio_path), sig)]),
            "MFCC": pd.DataFrame([export_mfcc_debug_row(str(audio_path), sig)]),
            "output_excel_path": str(output_excel_path),
        }
    except Exception as e:
        print(f" [ERROR] Failed to process {audio_path.name}. Reason: {e}")
        return None


def batch_process_directory(root_directory_path):
    root_dir = Path(root_directory_path)
    audio_files = [fp for fp in root_dir.rglob("*") if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTENSIONS]

    if not audio_files:
        print("No audio targets detected.")
        return

    print(f"Found {len(audio_files)} audio files. Launching parallel evaluation pipeline...")
    max_workers = max(1, (os.cpu_count() or 2) - 1)

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(process_audio_file, audio_files))

    master_sheets = {"Time": [], "Frequency": [], "Chroma": [], "Tempo": [], "MFCC": []}
    for res in results:
        if res is None:
            continue
        for sheet_name in master_sheets:
            master_sheets[sheet_name].append(res[sheet_name])

    master_excel_path = root_dir / "MASTER_AUDIO_DEBUG_SUMMARY.xlsx"
    with pd.ExcelWriter(master_excel_path, engine="openpyxl") as writer:
        index_rows = []
        for sheet_name, df_list in master_sheets.items():
            if df_list:
                combined_df = pd.concat(df_list, ignore_index=True)
            else:
                combined_df = pd.DataFrame()
            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            index_rows.append({
                "sheet_name": sheet_name,
                "rows": len(combined_df),
                "columns": len(combined_df.columns),
                "status_counts": combined_df["status"].value_counts().to_dict() if "status" in combined_df.columns else {},
            })
        pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)

    stylize_excel_file(master_excel_path)
    print(f"[SUCCESS] Consolidated Master Debug Sheet generated at:\n -> {master_excel_path.resolve()}")


if __name__ == "__main__":
    audio_root_path = r"D:/Engineering/Signal Processing/Personal Projects/Song Analysis/dataset/songs/"
    batch_process_directory(audio_root_path)